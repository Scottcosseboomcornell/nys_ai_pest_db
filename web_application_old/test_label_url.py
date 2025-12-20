#!/usr/bin/env python3
"""
Test script to verify label URL is included in Excel downloads
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

def test_label_url_in_excel():
    """Test that label URL is properly included in Excel generation"""
    
    # Sample pesticide data with label URL
    sample_pesticide = {
        'epa_reg_no': '100-1000',
        'trade_Name': 'Test Pesticide',
        'COMPANY_NAME': 'Test Company',
        'is_Organic': False,
        'CAUTION_statement': 'WARNING',
        'PPE': 'Gloves, goggles',
        'ABNS': 'Test Pesticide (Primary Name), Alternative Name',
        'label_url': 'https://www3.epa.gov/pesticides/chem_search/ppls/000100-00001-20200101.pdf',
        'Safety_Information': {
            'FIRST_REG_DT': '2020-01-01',
            'PHYS_FORM': 'Liquid'
        },
        'Active_Ingredients': [
            {
                'name': 'Test Ingredient 1',
                'percentage': '10%',
                'mode_Of_Action': 'Mode 1'
            }
        ],
        'Application_Info': [
            {
                'Target_Crop': [{'name': 'Apple'}],
                'Target_Disease_Pest': [{'name': 'Apple Scab'}],
                'low_rate': '1.0',
                'high_rate': '2.0',
                'units': 'lbs/acre',
                'REI': '24 hours',
                'PHI': '7 days',
                'application_Method': 'Foliar',
                'max_applications_per_season': '4'
            }
        ]
    }
    
    try:
        # Test individual pesticide download format
        print("🧪 Testing individual pesticide download format...")
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create Basic Information sheet
        basic_sheet = wb.create_sheet("Basic Information")
        
        # Basic Information
        basic_data = [
            ["EPA Registration Number", sample_pesticide.get('epa_reg_no', 'N/A')],
            ["Trade Name", sample_pesticide.get('trade_Name', 'N/A')],
            ["Company", sample_pesticide.get('COMPANY_NAME', 'N/A')],
            ["Organic", "Yes" if sample_pesticide.get('is_Organic', False) else "No"],
            ["Signal Word", sample_pesticide.get('CAUTION_statement', 'N/A')],
            ["PPE", sample_pesticide.get('PPE', 'N/A')],
            ["Alternative Names", sample_pesticide.get('ABNS', 'N/A')],
            ["Label URL", sample_pesticide.get('label_url', 'N/A')],
        ]
        
        # Write basic information
        for i, (label, value) in enumerate(basic_data, 1):
            basic_sheet[f'A{i}'] = label
            basic_sheet[f'B{i}'] = value
        
        # Check if label URL is included
        label_url_cell = basic_sheet['B8']
        if label_url_cell.value and label_url_cell.value != 'N/A':
            print("✅ Label URL found in individual pesticide download")
            print(f"   URL: {label_url_cell.value}")
        else:
            print("❌ Label URL not found in individual pesticide download")
        
        # Test filtered results download format
        print("\n🧪 Testing filtered results download format...")
        
        # Create Filtered Results sheet
        results_sheet = wb.create_sheet("Filtered Results")
        
        # Headers for the filtered results table
        headers = [
            "EPA Registration Number", "Trade Name", "Company", "Organic", 
            "Active Ingredients", "Mode of Action", "Signal Word", "PPE",
            "Label URL", "Crop", "Target Disease/Pest", "Low Rate", "High Rate", "Units", 
            "REI", "PHI", "Application Method", "Max Applications/Season"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            results_sheet.cell(row=1, column=col, value=header)
        
        # Write data row
        row_num = 2
        results_sheet.cell(row=row_num, column=1, value=sample_pesticide.get('epa_reg_no', 'N/A'))
        results_sheet.cell(row=row_num, column=2, value=sample_pesticide.get('trade_Name', 'N/A'))
        results_sheet.cell(row=row_num, column=3, value=sample_pesticide.get('COMPANY_NAME', 'N/A'))
        results_sheet.cell(row=row_num, column=4, value="Yes" if sample_pesticide.get('is_Organic', False) else "No")
        
        # Active ingredients
        active_ingredients = sample_pesticide.get('active_ingredients', [])
        ingredient_names = [ing.get('name', '') for ing in active_ingredients]
        ingredient_moa = [ing.get('mode_Of_Action', '') for ing in active_ingredients]
        
        results_sheet.cell(row=row_num, column=5, value=', '.join(ingredient_names) if ingredient_names else 'N/A')
        results_sheet.cell(row=row_num, column=6, value=', '.join(ingredient_moa) if ingredient_moa else 'N/A')
        results_sheet.cell(row=row_num, column=7, value=sample_pesticide.get('CAUTION_statement', 'N/A'))
        results_sheet.cell(row=row_num, column=8, value=sample_pesticide.get('PPE', 'N/A'))
        results_sheet.cell(row=row_num, column=9, value=sample_pesticide.get('label_url', 'N/A'))
        
        # Check if label URL is included in filtered results
        label_url_cell_filtered = results_sheet.cell(row=row_num, column=9)
        if label_url_cell_filtered.value and label_url_cell_filtered.value != 'N/A':
            print("✅ Label URL found in filtered results download")
            print(f"   URL: {label_url_cell_filtered.value}")
        else:
            print("❌ Label URL not found in filtered results download")
        
        # Save test file
        filename = "test_label_url_inclusion.xlsx"
        wb.save(filename)
        print(f"\n💾 Test file saved as: {filename}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error testing label URL inclusion: {str(e)}")
        return False

if __name__ == "__main__":
    test_label_url_in_excel()













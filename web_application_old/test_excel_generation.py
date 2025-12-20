#!/usr/bin/env python3
"""
Test script for Excel generation functionality
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

def test_excel_generation():
    """Test Excel generation with sample pesticide data"""
    
    # Sample pesticide data structure
    sample_pesticide = {
        'epa_reg_no': '100-1000',
        'trade_Name': 'Test Pesticide',
        'COMPANY_NAME': 'Test Company',
        'is_Organic': False,
        'CAUTION_statement': 'WARNING',
        'PPE': 'Gloves, goggles',
        'ABNS': 'Test Pesticide (Primary Name), Alternative Name',
        'Safety_Information': {
            'FIRST_REG_DT': '2020-01-01',
            'PHYS_FORM': 'Liquid'
        },
        'Active_Ingredients': [
            {
                'name': 'Test Ingredient 1',
                'percentage': '10%',
                'mode_Of_Action': 'Mode 1'
            },
            {
                'name': 'Test Ingredient 2',
                'percentage': '5%',
                'mode_Of_Action': 'Mode 2'
            }
        ],
        'Application_Info': [
            {
                'Target_Crop': [{'name': 'Apple'}],
                'Target_Disease_Pest': [{'name': 'Apple Scab'}],
                'low_rate': '1.0',
                'high_rate': '2.0',
                'units': 'lbs/acre',
                'REI': '24 hours',
                'PHI': '7 days',
                'application_Method': 'Foliar',
                'max_applications_per_season': '4'
            }
        ]
    }
    
    try:
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Basic Information sheet
        basic_sheet = wb.create_sheet("Basic Information")
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Basic Information
        basic_data = [
            ["EPA Registration Number", sample_pesticide.get('epa_reg_no', 'N/A')],
            ["Trade Name", sample_pesticide.get('trade_Name', 'N/A')],
            ["Company", sample_pesticide.get('COMPANY_NAME', 'N/A')],
            ["Organic", "Yes" if sample_pesticide.get('is_Organic', False) else "No"],
            ["Signal Word", sample_pesticide.get('CAUTION_statement', 'N/A')],
            ["PPE", sample_pesticide.get('PPE', 'N/A')],
            ["Alternative Names", sample_pesticide.get('ABNS', 'N/A')],
        ]
        
        # Add Safety Information if available
        if sample_pesticide.get('Safety_Information'):
            safety_info = sample_pesticide['Safety_Information']
            if safety_info.get('FIRST_REG_DT'):
                basic_data.append(["First Registration Date", safety_info['FIRST_REG_DT']])
            if safety_info.get('PHYS_FORM'):
                basic_data.append(["Physical Form", safety_info['PHYS_FORM']])
        
        # Write basic information
        for i, (label, value) in enumerate(basic_data, 1):
            basic_sheet[f'A{i}'] = label
            basic_sheet[f'B{i}'] = value
            
            # Style header
            basic_sheet[f'A{i}'].font = header_font
            basic_sheet[f'A{i}'].fill = header_fill
            basic_sheet[f'A{i}'].alignment = header_alignment
        
        # Set column widths
        basic_sheet.column_dimensions['A'].width = 25
        basic_sheet.column_dimensions['B'].width = 40
        
        # Create Active Ingredients sheet
        if sample_pesticide.get('Active_Ingredients'):
            ai_sheet = wb.create_sheet("Active Ingredients")
            
            # Headers for Active Ingredients
            ai_headers = ["Name", "Percentage", "Mode of Action"]
            for col, header in enumerate(ai_headers, 1):
                cell = ai_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add active ingredients data
            for row, ingredient in enumerate(sample_pesticide['Active_Ingredients'], 2):
                ai_sheet.cell(row=row, column=1, value=ingredient.get('name', 'N/A'))
                ai_sheet.cell(row=row, column=2, value=ingredient.get('percentage', 'N/A'))
                ai_sheet.cell(row=row, column=3, value=ingredient.get('mode_Of_Action', 'N/A'))
            
            # Set column widths
            ai_sheet.column_dimensions['A'].width = 30
            ai_sheet.column_dimensions['B'].width = 15
            ai_sheet.column_dimensions['C'].width = 25
        
        # Create Application Information sheet
        if sample_pesticide.get('Application_Info'):
            app_sheet = wb.create_sheet("Application Information")
            
            # Headers for Application Information
            app_headers = ["Crop", "Target Disease/Pest", "Low Rate", "High Rate", "Units", "REI", "PHI", "Application Method", "Max Applications/Season"]
            for col, header in enumerate(app_headers, 1):
                cell = app_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add application data
            for row, app in enumerate(sample_pesticide['Application_Info'], 2):
                # Extract crop names
                crops = [crop.get('name', '') for crop in app.get('Target_Crop', [])]
                crop_names = ', '.join(crops) if crops else 'N/A'
                
                # Extract disease/pest names
                diseases_pests = [dp.get('name', '') for dp in app.get('Target_Disease_Pest', [])]
                disease_pest_names = ', '.join(diseases_pests) if diseases_pests else 'N/A'
                
                app_sheet.cell(row=row, column=1, value=crop_names)
                app_sheet.cell(row=row, column=2, value=disease_pest_names)
                app_sheet.cell(row=row, column=3, value=app.get('low_rate', 'N/A'))
                app_sheet.cell(row=row, column=4, value=app.get('high_rate', 'N/A'))
                app_sheet.cell(row=row, column=5, value=app.get('units', 'N/A'))
                app_sheet.cell(row=row, column=6, value=app.get('REI', 'N/A'))
                app_sheet.cell(row=row, column=7, value=app.get('PHI', 'N/A'))
                app_sheet.cell(row=row, column=8, value=app.get('application_Method', 'N/A'))
                app_sheet.cell(row=row, column=9, value=app.get('max_applications_per_season', 'N/A'))
            
            # Set column widths
            for col in range(1, 10):
                app_sheet.column_dimensions[chr(64 + col)].width = 20
        
        # Save to file
        filename = f"test_excel_{sample_pesticide.get('trade_Name', 'pesticide')}_{sample_pesticide.get('epa_reg_no')}.xlsx"
        # Clean filename of invalid characters
        filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
        
        wb.save(filename)
        
        print(f"✅ Excel file generated successfully: {filename}")
        print(f"📊 Sheets created: {wb.sheetnames}")
        
        # Check file size
        file_size = os.path.getsize(filename)
        print(f"📁 File size: {file_size} bytes")
        
        return True
        
    except Exception as e:
        print(f"❌ Error generating Excel file: {str(e)}")
        return False

if __name__ == "__main__":
    test_excel_generation()



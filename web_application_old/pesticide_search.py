#!/usr/bin/env python3
"""
Pesticide Search Frontend - Unified Version
A web application to search through pesticide JSON files with automatic environment detection
"""

from flask import Flask, render_template, request, jsonify, send_from_directory, send_file
import json
import os
import glob
from pathlib import Path
import time
from functools import lru_cache
import threading
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import io
from pest_category_lookup import get_simplified_category, get_all_simplified_categories, get_categories_for_pesticide
from target_lookup import (
    get_simplified_target, get_simplified_targets_list, get_target_type, get_all_target_types, 
    get_simplified_targets_for_crop_and_type, get_original_targets_for_simplified_target
)

app = Flask(__name__)

# Custom Jinja2 filter for datetime formatting
@app.template_filter('datetime')
def datetime_filter(timestamp):
    """Format timestamp as readable date and time"""
    if timestamp is None:
        return "Unknown"
    try:
        from datetime import datetime
        dt = datetime.fromtimestamp(timestamp)
        return dt.strftime("%B %d, %Y at %I:%M %p")
    except Exception:
        return "Unknown"

# Environment detection and configuration
def detect_environment():
    """
    Detect whether we're running locally or on the server and return appropriate configuration.
    """
    # Check if we're running on the server by looking for the altered_json directory in the same folder
    current_dir = os.path.dirname(os.path.abspath(__file__))
    server_json_path = os.path.join(current_dir, "altered_json")
    local_json_path = os.path.join(current_dir, "..", "pipeline_critical_docs", "altered_json")
    
    # Check if the server_json_path exists and is NOT a symlink (indicating actual server deployment)
    if os.path.exists(server_json_path) and not os.path.islink(server_json_path):
        # We're on the server - data is in the same directory
        return {
            'json_dir': "altered_json",
            'cache_timeout': 3600,  # 1 hour cache for production
            'environment': 'server'
        }
    else:
        # We're running locally - data is in the parent directory
        return {
            'json_dir': "../pipeline_critical_docs/altered_json",
            'cache_timeout': 0,  # No cache for local development
            'environment': 'local'
        }

# Get environment-specific configuration
env_config = detect_environment()
OUTPUT_JSON_DIR = env_config['json_dir']
CACHE_TIMEOUT = env_config['cache_timeout']
ENVIRONMENT = env_config['environment']

# Allow overriding cache timeout via environment variable for local speedups
_override_timeout = os.environ.get('PEST_DB_CACHE_TIMEOUT')
if _override_timeout:
    try:
        CACHE_TIMEOUT = int(_override_timeout)
    except Exception:
        pass

# Configuration
SEARCH_RESULTS_LIMIT = 50

# Global cache and data structures
_pesticide_cache = {}
_cache_timestamp = 0
_search_index = defaultdict(list)  # Index for faster searching
_lock = threading.Lock()
_precomputed_index = None  # Loaded precomputed filter index if available
_epa_to_pesticide = {}  # Fast lookup by EPA reg number
_crops_stats_cache = None  # Cache for precomputed crops and stats

print(f"🌍 Environment detected: {ENVIRONMENT}")
print(f"📁 JSON directory: {OUTPUT_JSON_DIR}")
print(f"⏱️  Cache timeout: {CACHE_TIMEOUT} seconds ({'disabled' if CACHE_TIMEOUT == 0 else 'enabled'})")

def load_pesticide_data():
    """Load all pesticide data from JSON files with caching"""
    global _pesticide_cache, _cache_timestamp, _search_index
    
    current_time = time.time()
    
    # Check if cache is still valid
    if _pesticide_cache and (current_time - _cache_timestamp) < CACHE_TIMEOUT:
        return _pesticide_cache
    
    with _lock:
        # Double-check after acquiring lock
        if _pesticide_cache and (current_time - _cache_timestamp) < CACHE_TIMEOUT:
            return _pesticide_cache
        
        print("Loading pesticide data from JSON files...")
        start_time = time.time()
        
        pesticide_data = []
        _search_index.clear()
        
        # Get all JSON files in the output_json directory
        json_files = glob.glob(os.path.join(OUTPUT_JSON_DIR, "*.json"))
        
        for i, json_file in enumerate(json_files):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                    # Extract key information
                    if 'pesticide' in data:
                        pesticide = data['pesticide']
                        # Active ingredients as list of dicts with name, mode_Of_Action, and percentage
                        active_ingredients = []
                        for ing in pesticide.get('Active_Ingredients', []):
                            active_ingredients.append({
                                'name': ing.get('name', ''),
                                'mode_Of_Action': ing.get('mode_Of_Action', ing.get('mode_of_action', ing.get('mode_Of_Action', 'N/A'))),
                                'percentage': ing.get('percentage', 'N/A')
                            })
                        # Extract application information
                        application_info = []
                        for app in pesticide.get('Application_Info', []):
                            # Extract crop names
                            crops = [crop.get('name', '') for crop in app.get('Target_Crop', [])]
                            crop_names = ', '.join(crops) if crops else 'N/A'
                            
                            # Extract disease/pest names
                            diseases_pests = [dp.get('name', '') for dp in app.get('Target_Disease_Pest', [])]
                            disease_pest_names = ', '.join(diseases_pests) if diseases_pests else 'N/A'
                            
                            application_info.append({
                                'Target_Crop': crop_names,
                                'Target_Disease_Pest': disease_pest_names,
                                'low_rate': app.get('low_rate', 'N/A'),
                                'high_rate': app.get('high_rate', 'N/A'),
                                'units': app.get('units', 'N/A'),
                                'REI': app.get('REI', 'N/A'),
                                'PHI': app.get('PHI', 'N/A'),
                                'application_Method': app.get('application_Method', 'N/A'),
                                'max_applications_per_season': app.get('max_applications_per_season', 'N/A')
                            })

                        # Debug: Check if Safety_Information exists
                        if pesticide.get('epa_reg_no') == '100-1328':
                            print(f"DEBUG: Palladium JSON keys: {list(pesticide.keys())}")
                            print(f"DEBUG: Safety_Information: {pesticide.get('Safety_Information', 'MISSING')}")
                        
                        # Filter application info to only show allowed crops
                        filtered_application_info = _filter_application_info_for_allowed_crops(application_info)
                        
                        pesticide_info = {
                            'epa_reg_no': pesticide.get('epa_reg_no', 'N/A'),
                            'trade_Name': pesticide.get('trade_Name', 'N/A'),
                            'is_Organic': pesticide.get('is_Organic', False),
                            'label_url': pesticide.get('label_url', ''),
                            'active_ingredients': active_ingredients,
                            'application_info': filtered_application_info,
                            'PPE': pesticide.get('PPE', 'N/A'),
                            'CAUTION_statement': pesticide.get('CAUTION_statement', 'N/A'),
                            'COMPANY_NAME': pesticide.get('COMPANY_NAME', 'N/A'),
                            'ABNS': pesticide.get('ABNS', 'N/A'),
                            'Safety_Information': pesticide.get('Safety_Information', {}),
                            'filename': os.path.basename(json_file)
                        }
                        pesticide_data.append(pesticide_info)
                        
                        # Build search index for faster searching
                        _build_search_index(pesticide_info, len(pesticide_data) - 1)
                        
            except Exception as e:
                print(f"Error loading {json_file}: {e}")
                continue
        
        # Sort by trade name for consistent pagination
        pesticide_data.sort(key=lambda x: x['trade_Name'].lower())
        
        # Update cache
        _pesticide_cache = pesticide_data
        _cache_timestamp = current_time
        # Build EPA -> pesticide mapping for fast retrieval
        _epa_to_pesticide.clear()
        for p in _pesticide_cache:
            epa = p.get('epa_reg_no', '')
            if epa:
                _epa_to_pesticide[epa] = p
        
        load_time = time.time() - start_time
        print(f"Loaded {len(pesticide_data)} pesticide records in {load_time:.2f} seconds")
        
        return pesticide_data


def load_precomputed_index():
    """Load precomputed filter index if available."""
    global _precomputed_index
    if _precomputed_index is not None:
        return _precomputed_index
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(current_dir, "precomputed_filter_index.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                _precomputed_index = json.load(f)
                print("✅ Loaded precomputed_filter_index.json")
        else:
            _precomputed_index = None
    except Exception as e:
        print(f"Error loading precomputed_filter_index.json: {e}")
        _precomputed_index = None
    return _precomputed_index

def load_crops_stats():
    """Load precomputed crops and statistics if available"""
    global _crops_stats_cache
    if _crops_stats_cache is not None:
        return _crops_stats_cache
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(current_dir, "precomputed_crops_stats.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                _crops_stats_cache = json.load(f)
                print("✅ Loaded precomputed_crops_stats.json")
        else:
            _crops_stats_cache = None
    except Exception as e:
        print(f"Error loading precomputed_crops_stats.json: {e}")
        _crops_stats_cache = None
    return _crops_stats_cache

def _build_search_index(pesticide_info, index):
    """Build search index for faster searching"""
    # Index by EPA reg number
    epa_lower = pesticide_info['epa_reg_no'].lower()
    _search_index[f"epa:{epa_lower}"].append(index)
    
    # Index by trade name
    trade_lower = pesticide_info['trade_Name'].lower()
    _search_index[f"trade:{trade_lower}"].append(index)
    
    # Index by company name
    company_lower = pesticide_info['COMPANY_NAME'].lower()
    _search_index[f"company:{company_lower}"].append(index)
    
    # Index by active ingredients
    for ingredient in pesticide_info.get('active_ingredients', []):
        ing_name_lower = ingredient.get('name', '').lower()
        if ing_name_lower:
            _search_index[f"ingredient:{ing_name_lower}"].append(index)
    
    # Index by crops and pests
    for app in pesticide_info.get('application_info', []):
        crop_lower = app.get('Target_Crop', '').lower()
        if crop_lower and crop_lower != 'n/a':
            _search_index[f"crop:{crop_lower}"].append(index)
        
        pest_lower = app.get('Target_Disease_Pest', '').lower()
        if pest_lower and pest_lower != 'n/a':
            _search_index[f"pest:{pest_lower}"].append(index)

def search_pesticides_optimized(data, query, search_type):
    """Optimized search using index"""
    query = query.lower().strip()
    if not query:
        return []
    
    results = []
    seen_indices = set()
    
    if search_type == 'epa_reg_no':
        # Direct index lookup
        key = f"epa:{query}"
        if key in _search_index:
            for idx in _search_index[key]:
                if idx not in seen_indices:
                    results.append(data[idx])
                    seen_indices.add(idx)
    
    elif search_type == 'trade_Name':
        # Direct index lookup
        key = f"trade:{query}"
        if key in _search_index:
            for idx in _search_index[key]:
                if idx not in seen_indices:
                    results.append(data[idx])
                    seen_indices.add(idx)
    
    elif search_type == 'active_ingredient':
        # Search ingredient index
        key = f"ingredient:{query}"
        if key in _search_index:
            for idx in _search_index[key]:
                if idx not in seen_indices:
                    results.append(data[idx])
                    seen_indices.add(idx)
    
    elif search_type == 'crop':
        # Search crop index
        key = f"crop:{query}"
        if key in _search_index:
            for idx in _search_index[key]:
                if idx not in seen_indices:
                    results.append(data[idx])
                    seen_indices.add(idx)
    
    elif search_type == 'pest':
        # Search pest index
        key = f"pest:{query}"
        if key in _search_index:
            for idx in _search_index[key]:
                if idx not in seen_indices:
                    results.append(data[idx])
                    seen_indices.add(idx)
    
    elif search_type == 'both':
        # Search all indices
        search_keys = [
            f"epa:{query}",
            f"trade:{query}",
            f"company:{query}",
            f"ingredient:{query}",
            f"crop:{query}",
            f"pest:{query}"
        ]
        
        for key in search_keys:
            if key in _search_index:
                for idx in _search_index[key]:
                    if idx not in seen_indices:
                        results.append(data[idx])
                        seen_indices.add(idx)
    
    return results[:SEARCH_RESULTS_LIMIT]

def search_pesticides(data, query, search_type):
    """Relevance-ranked search over in-memory data with partial matching.
    Prioritizes trade name and ABNS; company matches have the lowest weight.
    Returns ALL matched records sorted by relevance (no hard cap).
    """
    query = query.lower().strip()
    if not query:
        return []

    def contains(text: str) -> bool:
        return query in (text or '').lower()

    def score_record(p: dict) -> float:
        score = 0.0
        trade = p.get('trade_Name', '')
        abns = p.get('ABNS', '')
        epa = p.get('epa_reg_no', '')
        company = p.get('COMPANY_NAME', '')

        if search_type in ('trade_Name', 'both'):
            tl = trade.lower()
            if tl == query:
                score += 10
            elif tl.startswith(query):
                score += 6
            elif query in tl:
                score += 4

        if search_type in ('both', 'abns', 'trade_Name'):
            al = (abns or '').lower()
            if query in al:
                score += 3

        if search_type in ('epa_reg_no', 'both'):
            if query == (epa or '').lower():
                score += 8
            elif contains(epa):
                score += 3

        if search_type in ('active_ingredient', 'both'):
            for ing in p.get('active_ingredients', []):
                if contains(ing.get('name', '')):
                    score += 2
                    break

        if search_type in ('crop', 'both'):
            for app in p.get('application_info', []):
                if contains(app.get('Target_Crop', '')):
                    score += 1.5
                    break

        if search_type in ('pest', 'both'):
            for app in p.get('application_info', []):
                if contains(app.get('Target_Disease_Pest', '')):
                    score += 1.5
                    break

        # Company name: extremely low priority; only add if nothing else matched
        if score == 0.0 and search_type in ('both',):
            if contains(company):
                score += 0.1

        return score

    # Score all records
    scored_results = []
    for p in data:
        score = score_record(p)
        if score > 0:
            scored_results.append((score, p))

    # Sort by score descending
    scored_results.sort(key=lambda x: x[0], reverse=True)
    return [p for _, p in scored_results]

@app.route('/')
def homepage():
    """Homepage"""
    return render_template('homepage.html')

@app.route('/pesticide-database')
def pesticide_database():
    """Pesticide database page"""
    last_updated = get_last_updated_timestamp()
    return render_template('search.html', last_updated=last_updated)

@app.route('/pesticide-database/info')
def pesticide_database_info():
    """Information page for pesticide database"""
    return render_template('info.html')

@app.route('/docs/<path:filename>')
def serve_docs(filename):
    """Serve documentation files"""
    import os
    docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'docs')
    file_path = os.path.join(docs_dir, filename)
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return f"File not found: {filename}", 404
    
    print(f"Serving file: {file_path}")
    response = send_from_directory(docs_dir, filename, as_attachment=True)
    return response

@app.route('/sgw')
def sgw():
    """SGW landing page"""
    return render_template('sgw_landing_page.html')

@app.route('/api/search')
def api_search():
    """API endpoint for searching pesticides"""
    query = request.args.get('q', '').strip()
    search_type = request.args.get('type', 'both')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 200, type=int)
    
    if not query:
        return jsonify({'pesticides': [], 'results': [], 'total': 0, 'query': '', 'pagination': {'page': 1, 'per_page': per_page, 'has_next': False, 'pages': 0}})
    
    # Load cached pesticide data
    pesticide_data = load_pesticide_data()
    
    # Search for matches using optimized search
    results = search_pesticides(pesticide_data, query, search_type)

    total = len(results)
    if per_page <= 0:
        per_page = 200
    start_idx = max((page - 1), 0) * per_page
    end_idx = start_idx + per_page
    current_page_data = results[start_idx:end_idx]
    has_next = end_idx < total
    
    return jsonify({
        'pesticides': current_page_data,
        'results': current_page_data,  # backward compatibility
        'total': total,
        'query': query,
        'search_type': search_type,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'has_next': has_next,
            'pages': (total + per_page - 1) // per_page
        }
    })

@app.route('/api/stats')
def api_stats():
    """API endpoint for getting statistics"""
    # Try to use precomputed data first
    crops_stats = load_crops_stats()
    if crops_stats and 'total_pesticides' in crops_stats and 'total_active_ingredients' in crops_stats:
        return jsonify({
            'total_pesticides': crops_stats['total_pesticides'],
            'unique_ingredients': crops_stats['total_active_ingredients']
        })
    
    # Fallback to dynamic computation
    pesticide_data = load_pesticide_data()

    # Use a set of (name, mode_Of_Action) tuples for uniqueness
    unique_ingredients = set()
    for p in pesticide_data:
        for ing in p['active_ingredients']:
            unique_ingredients.add((ing['name'], ing['mode_Of_Action']))

    stats = {
        'total_pesticides': len(pesticide_data),
        'unique_ingredients': len(unique_ingredients)
    }

    return jsonify(stats)

@app.route('/api/pesticides')
def api_pesticides():
    """API endpoint for getting all pesticides with pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    # Load cached pesticide data
    pesticide_data = load_pesticide_data()
    
    # Calculate pagination
    total = len(pesticide_data)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    
    # Get the current page of data
    current_page_data = pesticide_data[start_idx:end_idx]
    
    return jsonify({
        'pesticides': current_page_data,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'pages': (total + per_page - 1) // per_page,
            'has_next': end_idx < total,
            'has_prev': page > 1
        }
    })

@app.route('/api/pesticide/<epa_reg_no>')
def api_pesticide_detail(epa_reg_no):
    """API endpoint for getting detailed pesticide information including application data"""
    # Find the JSON file for this EPA registration number
    json_file = os.path.join(OUTPUT_JSON_DIR, f"{epa_reg_no}.json")
    
    if not os.path.exists(json_file):
        return jsonify({'error': 'Pesticide not found'}), 404
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        if 'pesticide' not in data:
            return jsonify({'error': 'Invalid data format'}), 400
            
        pesticide = data['pesticide']
        
        # Extract application information
        application_info = []
        for app in pesticide.get('Application_Info', []):
            # Extract crop names
            crops = [crop.get('name', '') for crop in app.get('Target_Crop', [])]
            crop_names = ', '.join(crops) if crops else 'N/A'
            
            # Extract disease/pest names
            diseases_pests = [dp.get('name', '') for dp in app.get('Target_Disease_Pest', [])]
            disease_pest_names = ', '.join(diseases_pests) if diseases_pests else 'N/A'
            
            application_info.append({
                'Target_Crop': crop_names,
                'Target_Disease_Pest': disease_pest_names,
                'low_rate': app.get('low_rate', 'N/A'),
                'high_rate': app.get('high_rate', 'N/A'),
                'units': app.get('units', 'N/A'),
                'REI': app.get('REI', 'N/A'),
                'PHI': app.get('PHI', 'N/A'),
                'application_Method': app.get('application_Method', 'N/A'),
                'max_applications_per_season': app.get('max_applications_per_season', 'N/A')
            })
        
        # Filter application info to only show allowed crops
        filtered_application_info = _filter_application_info_for_allowed_crops(application_info)
        
        return jsonify({
            'epa_reg_no': pesticide.get('epa_reg_no', ''),
            'trade_Name': pesticide.get('trade_Name', ''),
            'is_Organic': pesticide.get('is_Organic', False),
            'label_url': pesticide.get('label_url', ''),
            'active_ingredients': pesticide.get('Active_Ingredients', []),
            'PPE': pesticide.get('PPE', 'N/A'),
            'CAUTION_statement': pesticide.get('CAUTION_statement', 'N/A'),
            'COMPANY_NAME': pesticide.get('COMPANY_NAME', 'N/A'),
            'ABNS': pesticide.get('ABNS', 'N/A'),
            'Safety_Information': pesticide.get('Safety_Information', {}),
            'application_info': filtered_application_info
        })
        
    except Exception as e:
        return jsonify({'error': f'Error loading pesticide data: {str(e)}'}), 500

# Helper: normalize crop names for grouping
def _normalize_crop_name(crop_name):
    """Normalize crop name by removing parenthetical content and converting plurals to singular."""
    if not crop_name:
        return ''
    
    # Remove parenthetical content
    normalized = crop_name.split('(')[0].strip()
    
    # Handle special pluralization cases - updated to include all crops from the specified list
    plural_to_singular = {
        'Almonds': 'Almond',
        'Apricots': 'Apricot', 
        'Walnuts': 'Walnut',
        'Nectarines': 'Nectarine',
        'Oranges': 'Orange',
        'Grapes': 'Grape',
        'Beans': 'Bean',
        'Beets': 'Beet',
        'Potatoes': 'Potato',
        'Tomatoes': 'Tomato',
        'Peppers': 'Pepper',
        'Strawberries': 'Strawberry',
        'Blueberries': 'Blueberry',
        'Cranberries': 'Cranberry',
        'Raspberries': 'Raspberry',
        'Blackberries': 'Blackberry',
        'Cherries': 'Cherry',
        'Peaches': 'Peach',
        'Pears': 'Pear',
        'Pecans': 'Pecan',
        'Cucumbers': 'Cucumber',
        'Broccolis': 'Broccoli',
        'Spinaches': 'Spinach'
    }
    
    # Check exact matches first
    if normalized in plural_to_singular:
        normalized = plural_to_singular[normalized]
    # General rule: remove 's' from end if it ends with 's' and is longer than 3 characters
    elif len(normalized) > 3 and normalized.endswith('s') and not normalized.lower().endswith(('ss', 'us', 'is')):
        # Don't convert words that naturally end in 's'
        if not any(normalized.lower().endswith(ending) for ending in ['grass', 'citrus', 'asparagus']):
            normalized = normalized[:-1]
    
    return normalized

# Helper: check if a crop is a variant (plural/singular, different capitalization) of an allowed crop
def _is_crop_variant(crop, allowed_crop):
    """Check if a crop is a variant of an allowed crop (handles plurals and different capitalizations)."""
    crop_lower = crop.lower()
    allowed_lower = allowed_crop.lower()
    
    # Direct match (already handled above, but good to have)
    if crop_lower == allowed_lower:
        return True
    
    # Check if one is plural of the other
    if crop_lower == allowed_lower + 's':
        return True
    if allowed_lower == crop_lower + 's':
        return True
    
    # Handle special pluralization cases
    plural_mappings = {
        'almond': 'almonds',
        'apricot': 'apricots', 
        'walnut': 'walnuts',
        'nectarine': 'nectarines',
        'orange': 'oranges',
        'grape': 'grapes',
        'bean': 'beans',
        'beet': 'beets',
        'potato': 'potatoes',
        'tomato': 'tomatoes',
        'pepper': 'peppers',
        'strawberry': 'strawberries',
        'blueberry': 'blueberries',
        'cranberry': 'cranberries',
        'raspberry': 'raspberries',
        'blackberry': 'blackberries',
        'cherry': 'cherries'
    }
    
    # Check both directions of the mapping
    for singular, plural in plural_mappings.items():
        if (crop_lower == singular and allowed_lower == plural) or (crop_lower == plural and allowed_lower == singular):
            return True
    
    return False

# Helper: filter application info to only include allowed crops
def _filter_application_info_for_allowed_crops(application_info):
    """Filter application info to only show applications for allowed crops."""
    if not application_info:
        return []
    
    # Define the allowed crops from the AI prompt (from ai_main_o4_chat.py)
    allowed_crops = ["Apple", "Blackberry", "Blueberry", "Grape", "Cherry", "Cranberry", "Peach", "Pear", "Pecan",
                    "Strawberry", "Spinach", "Nectarine", "Orange", "Pepper", "Tomato", "Almond", "Apricot", 
                    "Potato", "Raspberry", "Walnut", "Cucumber", "Broccoli"]
    
    filtered_apps = []
    for app in application_info:
        crops_str = app.get('Target_Crop', '') or ''
        if not crops_str or crops_str == 'N/A':
            continue
            
        # Split comma-separated crops and check each one
        crops = [c.strip() for c in crops_str.split(',') if c.strip()]
        allowed_crops_in_app = []
        
        for crop in crops:
            # Check if this crop matches any of the allowed crops
            crop_matches = False
            for allowed_crop in allowed_crops:
                # Direct match (case-insensitive)
                if crop.lower() == allowed_crop.lower():
                    crop_matches = True
                    break
                # Check if crop is a plural/singular form of allowed crop
                if _is_crop_variant(crop, allowed_crop):
                    crop_matches = True
                    break
            
            if crop_matches:
                allowed_crops_in_app.append(crop)
        
        # Only include this application if it has at least one allowed crop
        if allowed_crops_in_app:
            # Create a new app entry with only the allowed crops
            filtered_app = app.copy()
            filtered_app['Target_Crop'] = ', '.join(allowed_crops_in_app)
            filtered_apps.append(filtered_app)
    
    return filtered_apps

# Helper: normalize pest names for grouping
def _normalize_pest_name(pest_name):
    """Normalize pest name by removing parenthetical content, life stages, and converting plurals to singular."""
    if not pest_name:
        return ''
    
    # Remove parenthetical content (life stages, notes, etc.)
    normalized = pest_name.split('(')[0].strip()
    
    # Handle capitalization consistency - convert to title case
    normalized = normalized.title()
    
    # Handle special pluralization cases for pest names
    plural_to_singular = {
        'Aphids': 'Aphid',
        'Mites': 'Mite', 
        'Maggots': 'Maggot',
        'Beetles': 'Beetle',
        'Weevils': 'Weevil',
        'Caterpillars': 'Caterpillar',
        'Leafrollers': 'Leafroller',
        'Leafminers': 'Leafminer',
        'Leafhoppers': 'Leafhopper',
        'Thrips': 'Thrips',  # Thrips is both singular and plural
        'Borers': 'Borer',
        'Scales': 'Scale',
        'Whiteflies': 'Whitefly',
        'Flies': 'Fly',
        'Moths': 'Moth',
        'Worms': 'Worm',
        'Bugs': 'Bug',
        'Ants': 'Ant',
        'Wasps': 'Wasp',
        'Nematodes': 'Nematode',
        'Grubs': 'Grub',
        'Loopers': 'Looper',
        'Armyworms': 'Armyworm',
        'Cutworms': 'Cutworm',
        'Webworms': 'Webworm',
        'Rootworms': 'Rootworm',
        'Earworms': 'Earworm',
        'Fruitworms': 'Fruitworm',
        'Hornworms': 'Hornworm',
        'Wireworms': 'Wireworm',
        'Grasshoppers': 'Grasshopper',
        'Crickets': 'Cricket'
    }
    
    # Check exact matches first
    if normalized in plural_to_singular:
        normalized = plural_to_singular[normalized]
    # General rule: remove 's' from end if it ends with 's' and is longer than 3 characters
    elif len(normalized) > 3 and normalized.endswith('s') and not normalized.lower().endswith(('ss', 'us', 'is')):
        # Don't convert words that naturally end in 's' or are already singular
        singular_endings = ['thrips', 'virus', 'fungus', 'citrus', 'asparagus', 'grass']
        if not any(normalized.lower().endswith(ending) for ending in singular_endings):
            normalized = normalized[:-1]
    
    return normalized

# Global cache for crop grouping
_crop_grouping_cache = {}
_crop_grouping_cache_timestamp = 0

def _get_crop_grouping():
    """Get crop grouping with caching."""
    global _crop_grouping_cache, _crop_grouping_cache_timestamp
    
    current_time = time.time()
    # Cache for 5 minutes
    if current_time - _crop_grouping_cache_timestamp < 300 and _crop_grouping_cache:
        return _crop_grouping_cache
    
    pesticide_data = load_pesticide_data()
    normalized_to_originals = {}  # normalized_name -> set of original names
    normalized_to_pests = {}      # normalized_name -> set of pests
    
    for p in pesticide_data:
        for app in p.get('application_info', []):
            crops_str = app.get('Target_Crop', '') or ''
            pests_str = app.get('Target_Disease_Pest', '') or ''
            
            # Split comma-separated strings, trim whitespace
            crops = [c.strip() for c in crops_str.split(',') if c.strip()]
            pests = [d.strip() for d in pests_str.split(',') if d.strip()]
            
            for crop in crops:
                normalized = _normalize_crop_name(crop)
                if normalized:
                    # Track original names for each normalized name
                    if normalized not in normalized_to_originals:
                        normalized_to_originals[normalized] = set()
                        normalized_to_pests[normalized] = set()
                    
                    normalized_to_originals[normalized].add(crop)
                    
                    # Add normalized pests for this normalized crop
                    for pest in pests:
                        normalized_pest = _normalize_pest_name(pest)
                        if normalized_pest:
                            normalized_to_pests[normalized].add(normalized_pest)
    
    # Convert sets to sorted lists
    result = {
        'normalized_crops': sorted(list(normalized_to_originals.keys())),
        'normalized_to_originals': {norm: sorted(list(origs)) for norm, origs in normalized_to_originals.items()},
        'normalized_to_pests': {norm: sorted(list(pests)) for norm, pests in normalized_to_pests.items()}
    }
    
    _crop_grouping_cache = result
    _crop_grouping_cache_timestamp = current_time
    return result

# Global cache for pest grouping
_pest_grouping_cache = {}
_pest_grouping_cache_timestamp = 0

def _get_pest_grouping():
    """Get pest grouping with caching."""
    global _pest_grouping_cache, _pest_grouping_cache_timestamp
    
    current_time = time.time()
    # Cache for 5 minutes
    if current_time - _pest_grouping_cache_timestamp < 300 and _pest_grouping_cache:
        return _pest_grouping_cache
    
    pesticide_data = load_pesticide_data()
    normalized_to_originals = {}  # normalized_pest_name -> set of original pest names
    
    for p in pesticide_data:
        for app in p.get('application_info', []):
            pests_str = app.get('Target_Disease_Pest', '') or ''
            
            # Split comma-separated strings, trim whitespace
            pests = [d.strip() for d in pests_str.split(',') if d.strip()]
            
            for pest in pests:
                normalized = _normalize_pest_name(pest)
                if normalized:
                    # Track original names for each normalized name
                    if normalized not in normalized_to_originals:
                        normalized_to_originals[normalized] = set()
                    
                    normalized_to_originals[normalized].add(pest)
    
    # Convert sets to sorted lists
    result = {
        'normalized_pests': sorted(list(normalized_to_originals.keys())),
        'normalized_to_originals': {norm: sorted(list(origs)) for norm, origs in normalized_to_originals.items()}
    }
    
    _pest_grouping_cache = result
    _pest_grouping_cache_timestamp = current_time
    return result

# Helper: get unique crops and pests mapping from cached data
def _compute_crops_and_pests():
    """Derive unique crops list and mapping crop -> set of pests from cached pesticide data."""
    grouping = _get_crop_grouping()
    
    # Define the allowed crops from the AI prompt (from ai_main_o4_chat.py)
    allowed_crops = ["Apple", "Blackberry", "Blueberry", "Grape", "Cherry", "Cranberry", "Peach", "Pear", "Pecan",
                    "Strawberry", "Spinach", "Nectarine", "Orange", "Pepper", "Tomato", "Almond", "Apricot", 
                    "Potato", "Raspberry", "Walnut", "Cucumber", "Broccoli"]
    
    # Filter crops to only include those in the allowed list (case-insensitive and handle plurals)
    filtered_crops = []
    filtered_crop_to_pests = {}
    
    for crop in grouping['normalized_crops']:
        # Check if this crop matches any of the allowed crops (case-insensitive)
        crop_matches = False
        for allowed_crop in allowed_crops:
            # Direct match (case-insensitive)
            if crop.lower() == allowed_crop.lower():
                crop_matches = True
                break
            # Check if crop is a plural/singular form of allowed crop
            if _is_crop_variant(crop, allowed_crop):
                crop_matches = True
                break
        
        if crop_matches:
            filtered_crops.append(crop)
            filtered_crop_to_pests[crop] = grouping['normalized_to_pests'].get(crop, set())
    
    return filtered_crops, filtered_crop_to_pests

@app.route('/api/enums/crops')
def api_enums_crops():
    """Get top 20 crops alphabetically sorted"""
    # Try to use precomputed data first
    crops_stats = load_crops_stats()
    if crops_stats and 'crops' in crops_stats:
        return jsonify({"crops": crops_stats['crops']})
    
    # Fallback to dynamic computation
    crops, _ = _compute_crops_and_pests()
    return jsonify({"crops": crops})

@app.route('/api/enums/pests')
def api_enums_pests():
    """Get all unique simplified targets for a given crop and target type."""
    crop = request.args.get('crop', '').strip()
    target_type = request.args.get('target_type', '').strip()
    
    # If precomputed available, use top targets for faster response
    pre = load_precomputed_index()
    if pre and crop and target_type:
        key = f"{crop}|{target_type}"
        entries = pre.get("top_targets", {}).get(key, [])
        # Format as "Name (count)"
        return jsonify({
            "crop": crop or None,
            "target_type": target_type or None,
            "pests": [f"{e['name']} ({e['count']})" for e in entries]
        })

    data = load_pesticide_data()
    
    if crop:
        # Build comprehensive mapping of all crop-target combinations
        all_crop_target_combinations = defaultdict(int)
        selected_crop_target_combinations = defaultdict(int)
        
        original_crop_names = [crop.lower()]
        # Add common variations
        if 'apple' in crop.lower():
            original_crop_names.extend(['apples', 'apple tree', 'apple trees'])
        elif 'grape' in crop.lower():
            original_crop_names.extend(['grapes', 'grapevine', 'grapevines'])
        # Add more variations as needed
        
        # Count all crop-target combinations and selected crop-target combinations
        for pesticide in data:
            for app in pesticide.get('application_info', []):
                crops = [c.strip().lower() for c in (app.get('Target_Crop', '') or '').split(',') if c.strip()]
                targets = [d.strip() for d in (app.get('Target_Disease_Pest', '') or '').split(',') if d.strip()]
                
                # Count all combinations
                for crop_name in crops:
                    for target in targets:
                        all_crop_target_combinations[(crop_name, target)] += 1
                
                # Count selected crop combinations
                if any(orig_crop in crops for orig_crop in original_crop_names):
                    for target in targets:
                        selected_crop_target_combinations[target] += 1
        
        # Calculate target relevance scores
        target_scores = {}
        for target, selected_count in selected_crop_target_combinations.items():
            # Calculate total occurrences of this target across all crops
            total_count = sum(count for (crop_name, target_name), count in all_crop_target_combinations.items() 
                            if target_name == target)
            
            # Calculate relevance score: how often this target appears with selected crop vs all crops
            if total_count > 0:
                relevance_score = selected_count / total_count
                target_scores[target] = {
                    'selected_count': selected_count,
                    'total_count': total_count,
                    'relevance_score': relevance_score
                }
        
        # Filter targets by relevance (only show targets that appear at least 20% of the time with selected crop)
        # or targets that appear at least 3 times with the selected crop
        relevant_targets = []
        for target, scores in target_scores.items():
            if scores['relevance_score'] >= 0.2 or scores['selected_count'] >= 3:
                relevant_targets.append(target)
        
        # Convert to simplified targets and filter by target type, counting unique pesticides
        simplified_target_pesticides = defaultdict(set)  # simplified_target -> set of pesticide epa_reg_nos
        for pesticide in data:
            for app in pesticide.get('application_info', []):
                crops = [c.strip().lower() for c in (app.get('Target_Crop', '') or '').split(',') if c.strip()]
                targets = [d.strip() for d in (app.get('Target_Disease_Pest', '') or '').split(',') if d.strip()]
                
                # Check if this application matches the selected crop
                if any(orig_crop in crops for orig_crop in original_crop_names):
                    for target in targets:
                        # Get list of simplified targets (handles comma-separated targets)
                        simplified_targets_list = get_simplified_targets_list(target)
                        target_type_for_target = get_target_type(target)
                        
                        # Filter by target type if specified
                        if target_type and target_type_for_target != target_type:
                            continue
                            
                        # Add this pesticide to each simplified target
                        for simplified_target in simplified_targets_list:
                            if simplified_target and simplified_target.strip():
                                simplified_target_pesticides[simplified_target].add(pesticide.get('epa_reg_no', ''))
        
        # Convert to counts
        simplified_target_counts = {target: len(pesticides) for target, pesticides in simplified_target_pesticides.items()}
    else:
        # If no crop specified, get all targets
        all_targets = set()
        for pesticide in data:
            for app in pesticide.get('application_info', []):
                targets = [d.strip() for d in (app.get('Target_Disease_Pest', '') or '').split(',') if d.strip()]
                for target in targets:
                    all_targets.add(target)
        
        # Convert to simplified targets and filter by target type, counting occurrences
        simplified_target_counts = defaultdict(int)
        for original_target in all_targets:
            # Get list of simplified targets (handles comma-separated targets)
            simplified_targets_list = get_simplified_targets_list(original_target)
            target_type_for_target = get_target_type(original_target)
            
            # Filter by target type if specified
            if target_type and target_type_for_target != target_type:
                continue
                
            # Add all simplified targets from the list with counts
            for simplified_target in simplified_targets_list:
                if simplified_target and simplified_target.strip():
                    simplified_target_counts[simplified_target] += 1
    
    # Create list of targets with counts
    pests_with_counts = []
    for target in sorted(simplified_target_counts.keys()):
        count = simplified_target_counts[target]
        pests_with_counts.append(f"{target} ({count})")
    
    return jsonify({
        "crop": crop or None, 
        "target_type": target_type or None,
        "pests": pests_with_counts
    })

@app.route('/api/enums/pest-categories')
def api_enums_pest_categories():
    """Get all unique simplified pest categories."""
    # Return the simplified categories from our lookup table
    simplified_categories = get_all_simplified_categories()
    return jsonify({"pest_categories": simplified_categories})

@app.route('/api/enums/target-types')
def api_enums_target_types():
    """Get all unique target types."""
    target_types = get_all_target_types()
    return jsonify({"target_types": target_types})

@app.route('/api/filter')
def api_filter_by_crop_pest():
    """Filter pesticides by selected crop, pest category, target type, and target with pagination."""
    crop = request.args.get('crop', '').strip()
    pest_cat = request.args.get('pest_cat', '').strip()
    target_type = request.args.get('target_type', '').strip()
    pest = request.args.get('pest', '').strip()
    page = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=50, type=int)
    if per_page <= 0:
        per_page = 50
    if per_page > 500:
        per_page = 500

    pre = load_precomputed_index()
    data = load_pesticide_data()
    
    # Get the original crop names that match the selected normalized crop
    original_crop_names = []
    if crop:
        grouping = _get_crop_grouping()
        original_crop_names = grouping['normalized_to_originals'].get(crop, [crop])
        # Convert to lowercase for matching
        original_crop_names = [c.lower() for c in original_crop_names]
    
    # Get the original pest names that match the selected simplified pest
    original_pest_names = []
    if pest:
        # Use the target lookup to get all original targets that map to this simplified target
        original_pest_names = get_original_targets_for_simplified_target(pest)
        # Convert to lowercase for matching
        original_pest_names = [p.lower() for p in original_pest_names]
    
    def matches_filter(p):
        """Check if pesticide matches crop, pest category, target type, and/or pest filter."""
        if not crop and not pest_cat and not target_type and not pest:
            return False
        
        # Check pest category filter first (applies to the whole pesticide, not individual applications)
        pest_cat_ok = True
        if pest_cat:
            pesticide_pest_cat = p.get('Safety_Information', {}).get('PEST_CAT', '')
            if pesticide_pest_cat and pesticide_pest_cat != '?' and pesticide_pest_cat != 'N/A':
                # Get simplified categories for this pesticide
                simplified_categories = get_categories_for_pesticide(pesticide_pest_cat)
                pest_cat_ok = pest_cat in simplified_categories
            else:
                pest_cat_ok = False
        
        if not pest_cat_ok:
            return False
        
        # Check crop, target type, and pest filters (apply to individual applications)
        for app in p.get('application_info', []):
            crops = [c.strip().lower() for c in (app.get('Target_Crop', '') or '').split(',') if c.strip()]
            pests = [d.strip().lower() for d in (app.get('Target_Disease_Pest', '') or '').split(',') if d.strip()]
            
            # Check if any of the original crop names match
            crop_ok = True if not crop else any(orig_crop in crops for orig_crop in original_crop_names)
            
            # Check target type filter
            target_type_ok = True
            if target_type:
                # Check if any of the pests in this application match the target type
                target_type_ok = False
                for pest_name in pests:
                    pest_target_type = get_target_type(pest_name)
                    if pest_target_type == target_type:
                        target_type_ok = True
                        break
            
            # Check if any of the original pest names match
            pest_ok = True if not pest else any(orig_pest in pests for orig_pest in original_pest_names)
            
            if crop_ok and target_type_ok and pest_ok:
                return True
        return False
    
    # Fast path: if precomputed list exists for crop/type/pest, use it
    if crop and target_type and pest and pre:
        key = f"{crop}|{target_type}|{pest}"
        epa_list = set(pre.get("lists", {}).get(key, []))
        if epa_list:
            # Use fast EPA mapping to avoid scanning entire dataset
            filtered_data = [ _epa_to_pesticide[epa] for epa in epa_list if epa in _epa_to_pesticide ]
        else:
            filtered_data = []
    else:
        # Fallback to dynamic filtering
        filtered_data = [p for p in data if matches_filter(p)]
    total = len(filtered_data)
    
    # Paginate
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    paginated_data = filtered_data[start_index:end_index]
    
    has_next = end_index < total
    
    return jsonify({
        'pesticides': paginated_data,
        'total': total,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'has_next': has_next,
            'total_pages': (total + per_page - 1) // per_page
        }
    })



@app.route('/api/last-updated')
def api_last_updated():
    """API endpoint for getting the last updated timestamp"""
    timestamp = get_last_updated_timestamp()
    if timestamp:
        from datetime import datetime
        dt = datetime.fromtimestamp(timestamp)
        return jsonify({
            'timestamp': timestamp,
            'formatted': dt.strftime("%B %d, %Y at %I:%M %p")
        })
    else:
        return jsonify({'timestamp': None, 'formatted': 'Unknown'})

@app.route('/api/cache/refresh')

def get_last_updated_timestamp():
    """Get the timestamp of the most recently modified JSON file"""
    try:
        json_files = glob.glob(os.path.join(OUTPUT_JSON_DIR, "*.json"))
        if not json_files:
            return None
        
        # Get the most recent modification time
        latest_time = max(os.path.getmtime(f) for f in json_files)
        return latest_time
    except Exception as e:
        print(f"Error getting last updated timestamp: {e}")
        return None

def refresh_cache():
    """API endpoint to manually refresh the cache"""
    global _pesticide_cache, _cache_timestamp, _search_index
    
    with _lock:
        _pesticide_cache = {}
        _cache_timestamp = 0
        _search_index.clear()
        
        # Force reload
        load_pesticide_data()
        
        return jsonify({
            'status': 'success',
            'message': 'Cache refreshed successfully',
            'total_pesticides': len(_pesticide_cache)
        })

@app.route('/api/filter/download')
def download_filtered_excel():
    """Download filtered pesticide results as Excel file."""
    crop = request.args.get('crop', '').strip()
    pest_cat = request.args.get('pest_cat', '').strip()
    target_type = request.args.get('target_type', '').strip()
    pest = request.args.get('pest', '').strip()
    
    # Use the same filtering logic as the main filter endpoint
    data = load_pesticide_data()
    
    # Get the original crop names that match the selected normalized crop
    original_crop_names = []
    if crop:
        grouping = _get_crop_grouping()
        original_crop_names = grouping['normalized_to_originals'].get(crop, [crop])
        # Convert to lowercase for matching
        original_crop_names = [c.lower() for c in original_crop_names]
    
    # Get the original pest names that match the selected simplified pest
    original_pest_names = []
    if pest:
        # Use the target lookup to get all original targets that map to this simplified target
        original_pest_names = get_original_targets_for_simplified_target(pest)
        # Convert to lowercase for matching
        original_pest_names = [p.lower() for p in original_pest_names]
    
    def matches_filter(p):
        """Check if pesticide matches crop, pest category, target type, and/or pest filter."""
        if not crop and not pest_cat and not target_type and not pest:
            return False
        
        # Check pest category filter first (applies to the whole pesticide, not individual applications)
        pest_cat_ok = True
        if pest_cat:
            pesticide_pest_cat = p.get('Safety_Information', {}).get('PEST_CAT', '')
            if pesticide_pest_cat and pesticide_pest_cat != '?' and pesticide_pest_cat != 'N/A':
                # Get simplified categories for this pesticide
                simplified_categories = get_categories_for_pesticide(pesticide_pest_cat)
                pest_cat_ok = pest_cat in simplified_categories
            else:
                pest_cat_ok = False
        
        if not pest_cat_ok:
            return False
        
        # Check crop, target type, and pest filters (apply to individual applications)
        for app in p.get('application_info', []):
            crops = [c.strip().lower() for c in (app.get('Target_Crop', '') or '').split(',') if c.strip()]
            pests = [d.strip().lower() for d in (app.get('Target_Disease_Pest', '') or '').split(',') if d.strip()]
            
            # Check if any of the original crop names match
            crop_ok = True if not crop else any(orig_crop in crops for orig_crop in original_crop_names)
            
            # Check target type filter
            target_type_ok = True
            if target_type:
                # Check if any of the pests in this application match the target type
                target_type_ok = False
                for pest_name in pests:
                    pest_target_type = get_target_type(pest_name)
                    if pest_target_type == target_type:
                        target_type_ok = True
                        break
            
            # Check if any of the original pest names match
            pest_ok = True if not pest else any(orig_pest in pests for orig_pest in original_pest_names)
            
            if crop_ok and target_type_ok and pest_ok:
                return True
        return False
    
    # Filter data
    filtered_data = [p for p in data if matches_filter(p)]
    
    if not filtered_data:
        return jsonify({'error': 'No results found for the specified filters'}), 404
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Pesticides"
    
    # Define headers
    headers = [
        "EPA Registration Number", "Trade Name", "Company", "Organic Status",
        "Active Ingredients", "Mode of Action", "Signal Word", "PPE",
        "Label URL", "Crop", "Target Disease/Pest", "Low Rate", "High Rate",
        "Units", "REI", "PHI", "Application Method", "Max Applications/Season",
        "Pest Category"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    row = 2
    for pesticide in filtered_data:
        # Get basic pesticide info
        epa_reg_no = pesticide.get('epa_reg_no', 'N/A')
        trade_name = pesticide.get('trade_Name', 'N/A')
        company = pesticide.get('COMPANY_NAME', 'N/A')
        is_organic = pesticide.get('is_Organic', False)
        signal_word = pesticide.get('CAUTION_statement', 'N/A')
        ppe = pesticide.get('PPE', 'N/A')
        label_url = pesticide.get('label_url', 'N/A')
        pest_category = pesticide.get('Safety_Information', {}).get('PEST_CAT', 'N/A')
        
        # Get active ingredients info
        active_ingredients = pesticide.get('Active_Ingredients', [])
        ingredient_names = []
        modes_of_action = []
        for ing in active_ingredients:
            ingredient_names.append(ing.get('name', 'N/A'))
            modes_of_action.append(ing.get('mode_Of_Action', 'N/A'))
        
        active_ingredients_str = '; '.join(ingredient_names) if ingredient_names else 'N/A'
        modes_of_action_str = '; '.join(modes_of_action) if modes_of_action else 'N/A'
        
        # Add application info rows
        application_info = pesticide.get('application_info', [])
        if application_info:
            for app in application_info:
                crop_name = app.get('Target_Crop', 'N/A')
                target_pest = app.get('Target_Disease_Pest', 'N/A')
                low_rate = app.get('low_rate', 'N/A')
                high_rate = app.get('high_rate', 'N/A')
                units = app.get('units', 'N/A')
                rei = app.get('REI', 'N/A')
                phi = app.get('PHI', 'N/A')
                app_method = app.get('application_Method', 'N/A')
                max_apps = app.get('max_applications_per_season', 'N/A')
                
                # Add row data
                row_data = [
                    epa_reg_no, trade_name, company, "Yes" if is_organic else "No",
                    active_ingredients_str, modes_of_action_str, signal_word, ppe,
                    label_url, crop_name, target_pest, low_rate, high_rate,
                    units, rei, phi, app_method, max_apps, pest_category
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
        else:
            # No application info, add basic pesticide info
            row_data = [
                epa_reg_no, trade_name, company, "Yes" if is_organic else "No",
                active_ingredients_str, modes_of_action_str, signal_word, ppe,
                label_url, 'N/A', 'N/A', 'N/A', 'N/A',
                'N/A', 'N/A', 'N/A', 'N/A', 'N/A', pest_category
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # Generate filename
    filename_parts = ['filtered_pesticides']
    if crop:
        filename_parts.append(f'crop_{crop.replace(" ", "_")}')
    if pest_cat:
        filename_parts.append(f'pest_cat_{pest_cat.replace(" ", "_")}')
    if target_type:
        filename_parts.append(f'target_type_{target_type.replace(" ", "_")}')
    if pest:
        filename_parts.append(f'target_{pest.replace(" ", "_")}')
    
    filename = '_'.join(filename_parts) + '.xlsx'
    
    # Clean filename
    filename = ''.join(c for c in filename if c.isalnum() or c in '._-')
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    # Create templates directory if it doesn't exist
    templates_dir = Path('templates')
    templates_dir.mkdir(exist_ok=True)
    
    print("🚀 Starting Unified Pesticide Search Frontend...")
    print(f"🌍 Environment: {ENVIRONMENT}")
    print(f"📁 Loading data from: {OUTPUT_JSON_DIR}")
    print(f"⏱️  Cache: {'Disabled' if CACHE_TIMEOUT == 0 else f'{CACHE_TIMEOUT}s timeout'}")
    
    # Pre-load data on startup
    data = load_pesticide_data()
    print(f"✅ Pre-loaded {len(data)} pesticide records with search indexing")
    
    # Start the Flask app with environment-specific port
    port = 5001 if ENVIRONMENT == 'server' else 5001
    print(f"🌐 Starting on port {port}")
    
    # Add health check endpoint for monitoring
    @app.route('/health')
    def health_check():
        """Health check endpoint for monitoring"""
        try:
            # Quick check if data is loaded
            if len(_pesticide_cache) > 0:
                return jsonify({
                    'status': 'healthy',
                    'timestamp': time.time(),
                    'records_loaded': len(_pesticide_cache),
                    'environment': ENVIRONMENT
                }), 200
            else:
                return jsonify({
                    'status': 'unhealthy',
                    'error': 'No data loaded'
                }), 503
        except Exception as e:
            return jsonify({
                'status': 'unhealthy',
                'error': str(e)
            }), 503
    
    # Configure Flask for production
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True) 